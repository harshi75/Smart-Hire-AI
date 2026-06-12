import io
import csv
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional

from app.database.db import get_db
from app.models.models import User, Candidate, CandidateScore
from app.utils.auth import get_current_user

router = APIRouter()


@router.get("/export")
def export_candidates(
    format: str = Query("csv", enum=["csv", "excel"]),
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(Candidate).filter(Candidate.owner_id == current_user.id)
    if status:
        q = q.filter(Candidate.status == status)
    candidates = q.order_by(Candidate.created_at.desc()).all()

    rows = []
    for c in candidates:
        best_score = db.query(CandidateScore).filter(CandidateScore.candidate_id == c.id).order_by(CandidateScore.match_score.desc()).first()
        rows.append({
            "ID": c.id,
            "Name": c.name or "",
            "Email": c.email or "",
            "Phone": c.phone or "",
            "Skills": ", ".join(c.skills or []),
            "Status": c.status or "new",
            "Match Score": round(best_score.match_score, 1) if best_score else "",
            "Hiring Probability": f"{round(best_score.hiring_probability, 1)}%" if best_score else "",
            "Parse Confidence": f"{round(c.parse_confidence * 100)}%" if c.parse_confidence else "",
            "File": c.file_name or "",
            "Added": c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
        })

    if format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            format = "csv"

    if format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidates"

        headers = list(rows[0].keys()) if rows else ["ID", "Name", "Email"]
        header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 2)

        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=smarthire_candidates.xlsx"},
        )

    # CSV
    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.read().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=smarthire_candidates.csv"},
    )
